from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Any, Optional, Dict, Union, TypedDict


class CellData(TypedDict, total=False):
    value: Any
    color: str  # ARGB: 'FFRRGGBB' или RGB: 'RRGGBB'


Cell = Union[Any, CellData]


class XLSXHandler:
    def __init__(self, base_dir: str = "."):
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _normalize_color(color: str) -> str:
        """
        Приводим цвет к ARGB формату, который ожидает openpyxl.
        Принимаем:
          - 'RRGGBB'
          - 'FFRRGGBB'
          - '#RRGGBB'
        """
        c = color.strip().lstrip("#")
        if len(c) == 6:
            return "FF" + c.upper()
        if len(c) == 8:
            return c.upper()
        raise ValueError(f"Некорректный color: {color}. Жду RRGGBB или FFRRGGBB (можно с #).")

    def create_from_list(
        self,
        data: List[List[Cell]],
        file_path: str,
        headers: Optional[List[str]] = None,
        column_widths: Optional[Dict[int, float]] = None,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        file_path = Path(file_path)

        # Заголовки
        start_row_idx = 1
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 18
            start_row_idx = 2

        # Данные: пишем построчно, чтобы можно было применять стиль к отдельным ячейкам
        for r, row in enumerate(data, start=start_row_idx):
            for c, item in enumerate(row, start=1):
                # item может быть обычным значением или dict {value, color}
                if isinstance(item, dict) and ("value" in item or "color" in item):
                    value = item.get("value")
                    color = item.get("color")
                else:
                    value = item
                    color = None

                cell = ws.cell(row=r, column=c, value=value)

                if color:
                    fill_color = self._normalize_color(color)
                    cell.fill = PatternFill(
                        fill_type="solid",
                        start_color=fill_color,
                        end_color=fill_color,
                    )

        # Ширины колонок (только явно заданные)
        column_widths = column_widths or {}
        for col_idx, width in column_widths.items():
            # col_idx у тебя 0-based, а в Excel 1-based
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = width

        wb.save(file_path)
        return file_path
